#!/usr/bin/env python3
"""
BTDM File Move/Copy/Convert Utility (single file or glob)

Choose exactly one of:
  --source <file|pattern>    # a single file OR a glob pattern
  --source-glob <pattern>    # explicit glob

Destination:
  --dest <dir-or-file>       # if multiple sources, dest must be a directory

Other options:
  --mode move|copy           # default: move
  --append-date yes|no       # default: no
  --date-format %Y%m%d       # default: %Y%m%d
  --skip-missing yes|no      # default: yes
  --skip-unchanged yes|no   # default: no (copy mode only; skip unchanged dest)

Conversion options:
  --convert-to csv|xlsx      # convert format while copying/moving
  --sheet-name Sheet1        # worksheet name when writing .xlsx (default: Sheet1)
  --csv-encoding utf-8       # encoding used when reading/writing CSV (default: utf-8)

Behaviors:
- Relative paths are resolved under BTDM_ROOT (env var if set, else script-relative root).
- With globbing (either --source contains *,?,[] or --source-glob used), multiple matches
  require --dest to be a directory.
- Destination itself is never treated as a glob.

Exit codes:
 0 success; 2 bad args; 3 no matches (and skip-missing=no); 4 dest must be directory; 5 partial failures
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable, List

# ---------------------------------------------------------------------------
# BTDM paths
# ---------------------------------------------------------------------------
_env_root = os.environ.get("BTDM_ROOT", "").strip()
if _env_root:
    BTDM_ROOT = os.path.abspath(_env_root)
else:
    BTDM_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
TABLES_DIR = os.path.join(BTDM_ROOT, "var", "tables")
YAML_PATH = os.path.join(BTDM_ROOT, "usr", "config", "default", "data_source_analysis.yaml")
OUTPUT_DIR = os.path.join(BTDM_ROOT, "var", "analysis")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def get_logger(level: int = logging.INFO) -> logging.Logger:
    logger = logging.getLogger("file_move")
    if logger.handlers:
        return logger
    logger.setLevel(level)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(asctime)s,%(levelname)s,%(name)s,%(message)s"))
    logger.addHandler(ch)
    return logger

logger = get_logger()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def resolve_path(p: str) -> Path:
    if not p:
        return Path(BTDM_ROOT)
    p = os.path.expanduser(p)
    pp = Path(p)
    if pp.is_absolute():
        return pp.resolve()
    return (Path(BTDM_ROOT) / pp).resolve()


def ensure_dir_for_target(target: Path) -> None:
    if target.suffix:
        target.parent.mkdir(parents=True, exist_ok=True)
    else:
        target.mkdir(parents=True, exist_ok=True)


def add_date_suffix(p: Path, date_format: str) -> Path:
    stamp = datetime.now().strftime(date_format)
    if p.suffix:
        return p.with_name(f"{p.stem}_{stamp}{p.suffix}")
    return p.with_name(f"{p.name}_{stamp}")


def files_look_unchanged(src: Path, dst: Path) -> bool:
    """
    Return True if destination file appears to be the same as source
    based on size and modification time.
    """
    try:
        s1 = src.stat()
        s2 = dst.stat()
    except FileNotFoundError:
        return False
    return (s1.st_size == s2.st_size) and (int(s1.st_mtime) == int(s2.st_mtime))


# -----------------------------
# Conversion helpers
# -----------------------------

def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception:
        logger.error("openpyxl is required for CSV↔XLSX conversions. Install in venv: pip install openpyxl")
        raise


def convert_csv_to_xlsx(src: Path, target: Path, sheet_name: str, csv_encoding: str) -> None:
    _require_openpyxl()
    from openpyxl import Workbook  # type: ignore
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    with src.open("r", newline="", encoding=csv_encoding) as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    ensure_dir_for_target(target)
    wb.save(str(target))


def convert_xlsx_to_csv(src: Path, target: Path, sheet_name: str, csv_encoding: str) -> None:
    _require_openpyxl()
    from openpyxl import load_workbook  # type: ignore
    wb = load_workbook(str(src), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    ensure_dir_for_target(target)
    with target.open("w", newline="", encoding=csv_encoding) as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])

# ---------------------------------------------------------------------------
# Core processing
# ---------------------------------------------------------------------------

def _looks_like_glob(text: str) -> bool:
    return any(ch in text for ch in "*?[]")


def collect_sources(source: str | None, source_glob: str | None) -> List[Path]:
    if source and source_glob:
        raise ValueError("Choose exactly one of --source or --source-glob")
    if not source and not source_glob:
        raise ValueError("You must pass --source or --source-glob")

    if source_glob:
        gp = resolve_path(source_glob)
        return list(Path(gp.parent).glob(gp.name))

    assert source is not None
    if _looks_like_glob(source):
        gp = resolve_path(source)
        return list(Path(gp.parent).glob(gp.name))

    sp = resolve_path(source)
    if not sp.exists():
        raise FileNotFoundError(f"Source file not found: {sp}")
    return [sp]

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(description="Move/Copy/Convert files with optional date suffix and rename.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--source", help="Path to a single source file (BTDM-relative or absolute). May include glob.")
    src.add_argument("--source-glob", help="Glob pattern for multiple source files")
    ap.add_argument("--dest", required=True, help="Destination directory or file path")
    ap.add_argument("--mode", choices=["move", "copy"], default="move")
    ap.add_argument("--append-date", choices=["yes", "no"], default="no")
    ap.add_argument("--date-format", default="%Y%m%d")
    ap.add_argument("--skip-missing", choices=["yes", "no"], default="yes")
    ap.add_argument("--skip-unchanged", choices=["yes", "no"], default="no", help="When mode=copy, skip if destination appears unchanged")
    ap.add_argument("--convert-to", choices=["csv", "xlsx"], help="Convert file format during transfer")
    ap.add_argument("--sheet-name", default="Sheet1", help="Worksheet name for XLSX read/write")
    ap.add_argument("--csv-encoding", default="utf-8", help="CSV encoding to read/write (default utf-8)")
    return ap


def main(argv: Iterable[str] | None = None) -> int:
    if argv is None:
        argv = sys.argv[1:]
    ap = build_arg_parser()
    args = ap.parse_args(argv)

    try:
        sources = collect_sources(args.source, args.source_glob)
    except FileNotFoundError as e:
        logger.error(str(e))
        return 2
    except ValueError as e:
        logger.error(str(e))
        return 2

    if not sources:
        if args.skip_missing == "yes":
            logger.info("No matching source files. (skip-missing=yes)")
            return 0
        else:
            logger.error("No matching source files.")
            return 3

    dest_input = resolve_path(args.dest)
    dest_is_dir = (dest_input.suffix == "") or dest_input.is_dir() or str(args.dest).endswith(("/", "\\"))

    if len(sources) > 1 and not dest_is_dir:
        logger.error("When multiple sources match, --dest must be a directory.")
        return 4

    ensure_dir_for_target(dest_input)

    ok = 0
    for src in sources:
        try:
            if dest_is_dir:
                target = dest_input / src.name
            else:
                target = dest_input

            if args.append_date == "yes":
                target = add_date_suffix(target, args.date_format)

            if args.convert_to:
                if args.convert_to == "xlsx":
                    target = target.with_suffix(".xlsx")
                    logger.info(f"Converting (CSV->XLSX): {src} -> {target}")
                    convert_csv_to_xlsx(src, target, args.sheet_name, args.csv_encoding)
                else:
                    target = target.with_suffix(".csv")
                    logger.info(f"Converting (XLSX->CSV): {src} -> {target}")
                    convert_xlsx_to_csv(src, target, args.sheet_name, args.csv_encoding)
                if args.mode == "move":
                    try:
                        src.unlink()
                        logger.info(f"Deleted source after convert (move mode): {src}")
                    except Exception as e:
                        logger.warning(f"Converted but could not delete source {src}: {e}")
            else:
                # Normal move/copy without format conversion
                ensure_dir_for_target(target)
                if args.mode == "copy" and args.skip_unchanged == "yes" and target.exists():
                    if files_look_unchanged(src, target):
                        logger.info(f"Skipping unchanged copy: {src} -> {target}")
                        ok += 1
                        continue
                if args.mode == "move":
                    logger.info(f"Moving: {src} -> {target}")
                    shutil.move(str(src), str(target))
                else:
                    logger.info(f"Copying: {src} -> {target}")
                    shutil.copy2(str(src), str(target))
            ok += 1
        except Exception as e:
            logger.exception(f"Failed to {args.mode} {src} -> {target}: {e}")

    logger.info(f"Completed: {ok}/{len(sources)} files processed.")
    return 0 if ok == len(sources) else 5

if __name__ == "__main__":
    sys.exit(main())