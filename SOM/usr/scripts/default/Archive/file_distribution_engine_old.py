#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
file_distribution_engine.py
----------------------------
Purpose:
  Base + Overlay file distribution and merge engine for interactive Excel files,
  plus simple one-way file distribution for non-CSV/XLSX artifacts.

Inputs:
  - Config YAML (VTX universal template recommended)
  - Base input files defined in config

Outputs:
  - Interactive XLSX file with one sheet per input CSV (base update)
  - CSV extracts of interactive tabs under var/interactive/data (interactive update)
  - One-way file copies for non-CSV/XLSX distributions
"""

from __future__ import annotations

import argparse
import logging
import os
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import pandas as pd
except Exception as e:
    raise SystemExit("Missing dependency: pandas. Install it in the VTX venv.") from e

try:
    import yaml  # PyYAML
except Exception as e:
    raise SystemExit("Missing dependency: pyyaml. Install it in the VTX venv.") from e

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception as e:
    raise SystemExit("Missing dependency: openpyxl. Install it in the VTX venv.") from e


# ---------------------------------------------------------------------
# Globals / Paths (VTX style)
# ---------------------------------------------------------------------

def resolve_vtx_root() -> Path:
    """
    Determine VTX_ROOT in a cross-platform way.
    Priority:
      1) env VTX_ROOT
      2) env BTDM_ROOT (legacy)
      3) infer relative to this file: <VTX_ROOT>/usr/scripts/default/<script>.py
    """
    env_root = os.environ.get("VTX_ROOT") or os.environ.get("BTDM_ROOT")
    if env_root:
        return Path(env_root).expanduser().resolve()

    here = Path(__file__).resolve()
    inferred = here.parents[3]  # default/<script>.py -> scripts -> usr -> VTX_ROOT
    return inferred


VTX_ROOT = resolve_vtx_root()

DEFAULT_CONFIG_DIR = VTX_ROOT / "usr" / "config" / "run"
DEFAULT_CONFIG_PATH = DEFAULT_CONFIG_DIR / "file_distribution_engine.yaml"


def vtx_path(path_str: str | Path, *, must_exist: bool = False) -> Path:
    if isinstance(path_str, Path):
        p = path_str
    else:
        s = str(path_str).strip()
        s = os.path.expandvars(s)
        s = os.path.expanduser(s)
        s = s.replace("VTX_ROOT" + os.sep, str(VTX_ROOT) + os.sep)
        s = s.replace("BTDM_ROOT" + os.sep, str(VTX_ROOT) + os.sep)
        s = s.replace("VTX_ROOT/", str(VTX_ROOT) + "/")
        s = s.replace("BTDM_ROOT/", str(VTX_ROOT) + "/")
        p = Path(s)

    if not p.is_absolute():
        p = (VTX_ROOT / p).resolve()

    if must_exist and not p.exists():
        raise FileNotFoundError(f"Path does not exist: {p}")
    return p


# ---------------------------------------------------------------------
# Logging (VTX style)
# ---------------------------------------------------------------------

def get_logger(component: str) -> logging.Logger:
    lib_dir = VTX_ROOT / "usr" / "lib"
    if str(lib_dir) not in sys.path:
        sys.path.insert(0, str(lib_dir))

    try:
        import vtx_logging  # type: ignore

        return vtx_logging.get_logger(component=component)
    except Exception:
        logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
        return logging.getLogger(component)


logger = get_logger(component="file_distribution_engine")


# ---------------------------------------------------------------------
# Config loading helpers
# ---------------------------------------------------------------------

def load_yaml(path: Path) -> Dict[str, Any]:
    raw = path.read_text(encoding="utf-8")
    doc = yaml.safe_load(raw) or {}
    if not isinstance(doc, dict):
        raise ValueError(f"Config root must be a mapping/dict: {path}")

    if "config" in doc and isinstance(doc.get("config"), dict):
        cfg = doc["config"]
    else:
        cfg = doc

    for k in ("reports", "jobs", "stretches", "inputs", "outputs"):
        if k in cfg and cfg[k] is None:
            cfg[k] = []

    return cfg


# ---------------------------------------------------------------------
# CLI / Options
# ---------------------------------------------------------------------

@dataclass
class Options:
    config_path: Path
    job_id: Optional[str] = None
    dry_run: bool = False


def parse_args(argv: Optional[list[str]] = None) -> Options:
    p = argparse.ArgumentParser(description="Base + overlay file distribution engine")
    p.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG_PATH),
        help=f"Path to YAML config (default: {DEFAULT_CONFIG_PATH})",
    )
    p.add_argument("--job", default=None, help="Run only a single job id")
    p.add_argument("--dry-run", action="store_true", help="Do not write outputs")
    args = p.parse_args(argv)

    cfg_path = vtx_path(args.config, must_exist=True)
    return Options(config_path=cfg_path, job_id=args.job, dry_run=bool(args.dry_run))


# ---------------------------------------------------------------------
# Core helpers
# ---------------------------------------------------------------------

def read_csv_base(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def sheet_to_dataframe(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    header = ["" if v is None else str(v) for v in rows[0]]
    data_rows = rows[1:]

    cleaned = []
    for row in data_rows:
        if row is None:
            continue
        if all(cell is None for cell in row):
            continue
        cleaned.append(["" if v is None else str(v) for v in row])

    if not any(header):
        return pd.DataFrame()

    return pd.DataFrame(cleaned, columns=header)


def normalize_anchor(anchor: str, base_df: pd.DataFrame) -> str:
    if anchor.lower() == "first" or anchor.strip() == "":
        return str(base_df.columns[0])
    return anchor


def merge_base_overlay(base_df: pd.DataFrame, existing_df: pd.DataFrame, anchor_col: str) -> pd.DataFrame:
    if existing_df.empty or anchor_col not in existing_df.columns:
        return base_df.copy()

    overlay_cols = [c for c in existing_df.columns if c not in base_df.columns]
    if not overlay_cols:
        return base_df.copy()

    overlay_df = existing_df[[anchor_col] + overlay_cols].copy()
    overlay_df = overlay_df.drop_duplicates(subset=[anchor_col], keep="last")

    merged = base_df.merge(overlay_df, on=anchor_col, how="left")
    return merged


def sanitize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cleaned = df.copy()
    for col in cleaned.columns:
        cleaned[col] = cleaned[col].map(sanitize_cell)
    return cleaned


INVALID_SHEET_CHARS = re.compile(r"[:\\\\/?*\\[\\]]")


def sanitize_sheet_name(name: str) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("_", name).strip()
    cleaned = cleaned.strip("'")
    if not cleaned:
        cleaned = "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def make_unique_sheet_names(names: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for name in names:
        base = name
        if base not in seen:
            seen[base] = 1
            out.append(base)
            continue
        seen[base] += 1
        suffix = f"_{seen[base]}"
        trimmed = base
        if len(trimmed) + len(suffix) > 31:
            trimmed = trimmed[: 31 - len(suffix)]
        unique = f"{trimmed}{suffix}"
        while unique in seen:
            seen[base] += 1
            suffix = f"_{seen[base]}"
            trimmed = base
            if len(trimmed) + len(suffix) > 31:
                trimmed = trimmed[: 31 - len(suffix)]
            unique = f"{trimmed}{suffix}"
        seen[unique] = 1
        out.append(unique)
    return out


def update_sheet_in_place(ws, df: pd.DataFrame) -> None:
    safe_df = sanitize_dataframe(df)
    columns = list(safe_df.columns)
    values = safe_df.values.tolist()

    new_rows = 1 + len(values)
    new_cols = len(columns)

    max_row = max(ws.max_row, new_rows)
    max_col = max(ws.max_column, new_cols)

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for r_idx, row in enumerate(values, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for r_idx in range(new_rows + 1, max_row + 1):
        for c_idx in range(1, max_col + 1):
            ws.cell(row=r_idx, column=c_idx, value=None)

    for c_idx in range(new_cols + 1, max_col + 1):
        for r_idx in range(1, max_row + 1):
            ws.cell(row=r_idx, column=c_idx, value=None)


def write_new_sheet(ws, df: pd.DataFrame) -> None:
    safe_df = sanitize_dataframe(df)
    columns = list(safe_df.columns)
    values = safe_df.values.tolist()

    ws.append(columns)
    for row in values:
        ws.append(row)


def ensure_workbook(path: Path) -> Tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), True
    wb = Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])
    return wb, False


def decide_mode(inputs: List[Path], output: Path, forced_mode: Optional[str]) -> str:
    if forced_mode:
        return forced_mode
    if not output.exists():
        return "base"
    input_mtime = max(p.stat().st_mtime for p in inputs if p.exists())
    output_mtime = output.stat().st_mtime
    if output_mtime > input_mtime:
        return "interactive"
    return "base"


def is_dir_path(path: Path) -> bool:
    if path.exists():
        return path.is_dir()
    return path.suffix == ""


# ---------------------------------------------------------------------
# Job execution
# ---------------------------------------------------------------------

def run_base_update_excel(
    input_paths: List[Path],
    xlsx_path: Path,
    anchor_column: str,
    tab_names: Optional[List[str]],
    sanitize_tabs: bool,
    dry_run: bool,
) -> None:
    wb, _ = ensure_workbook(xlsx_path)
    derived_names: List[str] = []
    if tab_names is not None:
        if len(tab_names) != len(input_paths):
            raise ValueError("tabs list must match inputs length")
        derived_names = [str(t) for t in tab_names]
    else:
        derived_names = [p.stem for p in input_paths]

    if sanitize_tabs:
        derived_names = [sanitize_sheet_name(n) for n in derived_names]
    derived_names = make_unique_sheet_names(derived_names)

    for csv_path, tab_name in zip(input_paths, derived_names):
        if not csv_path.exists():
            raise FileNotFoundError(f"Input CSV does not exist: {csv_path}")

        base_df = read_csv_base(csv_path)
        if base_df.empty:
            logger.warning("Base CSV is empty: %s", csv_path)

        anchor = normalize_anchor(anchor_column, base_df)
        if anchor not in base_df.columns:
            raise ValueError(f"Anchor column '{anchor}' not found in base CSV: {csv_path}")

        if tab_name in wb.sheetnames:
            ws = wb[tab_name]
            existing_df = sheet_to_dataframe(ws)
            merged = merge_base_overlay(base_df, existing_df, anchor)
            update_sheet_in_place(ws, merged)
        else:
            ws = wb.create_sheet(title=tab_name)
            write_new_sheet(ws, base_df)

    if dry_run:
        logger.info("Dry run: skipping workbook write (%s)", xlsx_path)
        return

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    logger.info("Workbook updated: %s", xlsx_path)


def run_base_copy(inputs: List[Path], outputs: List[Path], dry_run: bool) -> None:
    if len(outputs) == 1 and len(inputs) == 1:
        pairs = [(inputs[0], outputs[0])]
    elif len(outputs) == len(inputs):
        pairs = list(zip(inputs, outputs))
    else:
        raise ValueError("Base copy requires outputs length 1 or equal to inputs length")

    for src, dest in pairs:
        if not src.exists():
            raise FileNotFoundError(f"Input does not exist: {src}")
        if dry_run:
            logger.info("Dry run: skipping copy %s -> %s", src, dest)
            continue
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest)
        logger.info("Copied %s -> %s", src, dest)


def run_interactive_update(
    xlsx_path: Path,
    outputs: List[Path],
    tabs: Optional[List[str]],
    sanitize_tabs: bool,
    dry_run: bool,
) -> None:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Interactive XLSX does not exist: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)

    tab_names = [str(t) for t in tabs] if isinstance(tabs, list) else None
    if tab_names is not None:
        if sanitize_tabs:
            tab_names = [sanitize_sheet_name(n) for n in tab_names]
        tab_names = make_unique_sheet_names(tab_names)

    if len(outputs) == 1 and is_dir_path(outputs[0]):
        output_dir = outputs[0]
        output_dir.mkdir(parents=True, exist_ok=True)
        target_tabs = tab_names or list(wb.sheetnames)
        for tab_name in target_tabs:
            if tab_name not in wb.sheetnames:
                logger.warning("Tab not found in workbook: %s", tab_name)
                continue
            ws = wb[tab_name]
            df = sheet_to_dataframe(ws)
            output_path = output_dir / f"{tab_name}.csv"
            if dry_run:
                logger.info("Dry run: skipping write %s", output_path)
                continue
            df.to_csv(output_path, index=False)
            logger.info("Wrote interactive CSV: %s", output_path)
        return

    output_map = {p.stem: p for p in outputs}
    target_tabs = tab_names or list(output_map.keys())
    for tab_name in target_tabs:
        if tab_name not in wb.sheetnames:
            logger.warning("Tab not found in workbook: %s", tab_name)
            continue
        if tab_name not in output_map:
            logger.warning("No output mapping for tab: %s", tab_name)
            continue
        ws = wb[tab_name]
        df = sheet_to_dataframe(ws)
        output_path = output_map[tab_name]
        if dry_run:
            logger.info("Dry run: skipping write %s", output_path)
            continue
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output_path, index=False)
        logger.info("Wrote interactive CSV: %s", output_path)


# ---------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------

def main(argv: Optional[list[str]] = None) -> int:
    opt = parse_args(argv)
    logger.info("VTX_ROOT=%s", VTX_ROOT)
    logger.info("Config=%s", opt.config_path)

    cfg = load_yaml(opt.config_path)
    payload = cfg.get("payload", {}) if isinstance(cfg.get("payload"), dict) else {}
    jobs = payload.get("jobs", []) or []

    if not isinstance(jobs, list):
        raise ValueError("config.payload.jobs must be a list")

    selected_jobs = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        if not job.get("enabled", True):
            continue
        if opt.job_id and str(job.get("id")) != str(opt.job_id):
            continue
        selected_jobs.append(job)

    if not selected_jobs:
        logger.warning("No jobs to run")
        return 0

    for job in selected_jobs:
        job_id = str(job.get("id", ""))
        logger.info("Running job: %s", job_id)

        inputs_raw = job.get("inputs", []) or []
        outputs_raw = job.get("outputs", []) or []

        if not inputs_raw:
            raise ValueError(f"Job '{job_id}' has no inputs")
        if not outputs_raw:
            raise ValueError(f"Job '{job_id}' has no outputs")

        inputs = [vtx_path(p, must_exist=False) for p in inputs_raw]
        outputs = [vtx_path(p, must_exist=False) for p in outputs_raw]

        mode = str(job.get("mode", "base")).lower()
        if mode not in ("auto", "base", "interactive"):
            raise ValueError(f"Job '{job_id}' mode must be auto|base|interactive")

        anchor_column = str(job.get("anchor_column", "first"))
        tabs = job.get("tabs") or job.get("tab_names")
        tabs = [str(t) for t in tabs] if isinstance(tabs, list) else None
        sanitize_tabs = bool(job.get("sanitize_tabs", True))

        output_xlsx = len(outputs) == 1 and outputs[0].suffix.lower() == ".xlsx"
        inputs_csv = all(p.suffix.lower() == ".csv" for p in inputs)

        forced_mode = None if mode == "auto" else mode
        if mode == "auto":
            if output_xlsx:
                run_mode = decide_mode(inputs, outputs[0], forced_mode)
            else:
                run_mode = "base"
        else:
            run_mode = mode

        logger.info("Job '%s' mode: %s", job_id, run_mode)

        if run_mode == "base":
            if output_xlsx:
                if not inputs_csv:
                    raise ValueError(f"Job '{job_id}' base update requires CSV inputs for XLSX output")
                run_base_update_excel(inputs, outputs[0], anchor_column, tabs, sanitize_tabs, opt.dry_run)
            else:
                run_base_copy(inputs, outputs, opt.dry_run)
        else:
            if len(inputs) != 1 or inputs[0].suffix.lower() != ".xlsx":
                raise ValueError(f"Job '{job_id}' interactive update requires a single XLSX input")
            run_interactive_update(inputs[0], outputs, tabs, sanitize_tabs, opt.dry_run)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
